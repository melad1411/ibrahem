from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import json
from decimal import Decimal

# اسم ملف الإكسل
excel_file = "a.xlsx"

# نفتح ملف الإكسل
wb = load_workbook(excel_file, data_only=True)
ws = wb.active

# نقرأ رؤوس الأعمدة (الصف الأول)
headers = [cell.value for cell in ws[1]]

data = []
for row in ws.iter_rows(min_row=2):
    record = {}
    for idx, cell in enumerate(row):
        # هنا ناخد القيمة كما تُعرض (displayed) في إكسل
        val = cell.value
        if val is None:
            val_str = ""  # خلي الخلايا الفارغة نص فارغ
        else:
            # حول الرقم أو النص إلى string بالضبط كما يظهر
            val_str = str(val)
        record[headers[idx]] = val_str
    data.append(record)


# حفظ النتيجة كملف JSON
with open("data.json", "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=4)

print("Created JSON file!!")
